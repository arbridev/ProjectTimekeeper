## external imports

import openpyxl
import os
import string
import datetime
from openpyxl.styles import Font, Alignment

class ExcelHandler():

    workbook = openpyxl.Workbook()
    filepath = ''

    __alphabet = string.ascii_uppercase

    ## file methods

    def new_workbook(self, path):
        self.workbook = openpyxl.Workbook()
        self.filepath = path
        self.workbook.save(path)

    def open_workbook(self, path):
        self.workbook.close()
        self.workbook = openpyxl.load_workbook(filename = path)
        self.filepath = path

    def save_workbook(self):
        self.workbook.save(self.filepath)

    def close(self):
        self.workbook.close()

    ## input methods

    def input_text(self, text, cell):
        sheet = self.workbook.active
        sheet[cell] = text

    def input_number(self, stringNumber, cell):
        sheet = self.workbook.active
        sheet[cell].value = int(stringNumber.replace('.', ''))

    def input_date(self, date, cellCoordinate):
        sheet = self.workbook.active
        sheet[cellCoordinate].value = date

    def input_current_date(self, cell):
        self.input_date(datetime.datetime.now(), cell.coordinate)

    def input_scaffold(self):
        sheet = self.workbook.active
        header = [sheet['A1'], sheet['B1'], sheet['C1'], sheet['D1'], sheet['E1']]
        header[0].value = 'Id'
        header[1].value = 'Start'
        header[2].value = 'End'
        header[3].value = 'Task'
        header[4].value = 'Duration'
        self.adjust_column("B", 20)
        self.adjust_column("C", 20)
        self.adjust_column("D", 40)
        self.adjust_column("E", 15)
        self.adjust_column("F", 20)
        list(map(lambda c: self.apply_header(c), header))

    ## helper methods

    def get_next_empty_cell_down(self, column="A"):
        sheet = self.workbook.active
        nextrow = 1
        for cell in sheet[column]:
            if cell.value is None:
                nextrow = cell.row
                break
            else:
                nextrow = cell.row + 1
        return sheet[column + str(nextrow)]

    def get_last_filled_cell_vertically(self, column="A", from_row=1):
        sheet = self.workbook.active
        nextrow = 1
        for cell in sheet[column]:
            if cell.row < from_row:
                continue
            if cell.value is None:
                break
            else:
                nextrow = cell.row
        return sheet[column + str(nextrow)]

    def get_cell_above(self, cell):
        if cell.row - 1 > 0:
            return cell.offset(row = -1)
        else:
            return None

    def get_cell_below(self, cell):
        return cell.offset(row = 1)

    def get_cell_beside(self, cell, rightside=True):
        if rightside:
            return cell.offset(column = 1)
        else:
            if cell.column - 1 > 0:
                return cell.offset(column = -1)
            else:
                return None

    def adjust_column(self, column, width):
        sheet = self.workbook.active
        sheet.column_dimensions[column].width = width

    def adjust_row(self, row, height):
        sheet = self.workbook.active
        sheet.row_dimensions[row].height = height

    def insert_row(self, at_row):
        sheet = self.workbook.active
        sheet.insert_rows(at_row)
    
    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        sheet = self.workbook.active
        sheet.merge_cells(range_string=range_string, start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)

    def get_column(self, columncoord):
        sheet = self.workbook.active
        return sheet[columncoord]

    def get_cell(self, cellcoord):
        sheet = self.workbook.active
        return sheet[cellcoord]

    def apply_center(self, cell):
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def apply_font_bold(self, cell):
        cell.font = Font(bold=True)

    def apply_font_size(self, cell, size):
        cell.font = Font(size=size)

    def apply_font_color(self, cell, color):
        cell.font = Font(color=color)

    def apply_header(self, cell):
        cell.font = Font(name='Calibri', size=14, bold=True, italic=False,vertAlign=None, underline='none', strike=False, color='FF000000')
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def apply_font_style(self, cell, name='Calibri', size=11, bold=False, italic=False,vertAlign=None, underline='none', strike=False, color='FF000000'):
        cell.font = Font(name=name, size=size, bold=bold, italic=italic, vertAlign=vertAlign, underline=underline, strike=strike, color=color)